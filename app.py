import streamlit as st
import pandas as pd
import sqlite3
from datetime import datetime
import pytz
import urllib.parse
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

# ------------------ CONFIGURACIÓN GENERAL ------------------

with st.spinner('Iniciando sistema Champlitte... 🥐'):
    zona_mx = pytz.timezone('America/Mexico_City')
    fecha_hoy_mx = datetime.now(zona_mx).date()

st.set_page_config(page_title="Inventario Champlitte MX", page_icon="🥐", layout="wide")

# ------------------ WHATSAPP ------------------

contactos_whatsapp = {
    "Administrador": "522283530069",
    "Sucursal": "522299359597"
}

contacto_seleccionado = st.sidebar.selectbox(
    "📲 Enviar reportes a:",
    list(contactos_whatsapp.keys())
)

numero_whatsapp = contactos_whatsapp[contacto_seleccionado]

# Contenedores para mensajes
msg_conteo = st.empty()
msg_tabla = st.empty()
msg_corte = st.empty()

# ------------------ BASE DE DATOS ------------------

conn = sqlite3.connect('inventario_pan.db', check_same_thread=False)
c = conn.cursor()

c.execute('CREATE TABLE IF NOT EXISTS captura_actual (nombre TEXT, fecha_cad DATE, cantidad INTEGER)')
c.execute('CREATE TABLE IF NOT EXISTS base_anterior (nombre TEXT, fecha_cad DATE, cantidad INTEGER)')

c.execute('''CREATE TABLE IF NOT EXISTS historial_ventas (
    nombre TEXT,
    fecha_cad DATE,
    habia INTEGER,
    quedan INTEGER,
    vendidos INTEGER,
    fecha_corte DATETIME
)''')

conn.commit()

# ------------------ FUNCIONES ------------------

def sonido_click():
    st.markdown(
        """
        <audio autoplay>
        <source src="https://www.soundjay.com/buttons/sounds/button-16.mp3">
        </audio>
        """,
        unsafe_allow_html=True
    )

def sumar(valor):
    st.session_state.conteo_temp += valor
    sonido_click()

def resetear():
    st.session_state.conteo_temp = 0
    sonido_click()

def generar_csv(df):
    return df.to_csv(index=False).encode('utf-8')

def generar_excel_costa_verde(conn):

    df = pd.read_sql("SELECT nombre, fecha_cad, cantidad FROM base_anterior", conn)

    wb = Workbook()
    ws = wb.active
    ws.title = "COSTA VERDE"

    verde = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")

    ws.merge_cells("A1:I1")
    ws["A1"] = "COSTA VERDE"
    ws["A1"].font = Font(size=24, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    fila = 3

    ws["A3"] = "PRODUCTO"
    ws["B3"] = "CANTIDAD"
    ws["C3"] = "CADUCIDAD"

    for cell in ws["3:3"]:
        cell.font = Font(bold=True)
        cell.fill = verde

    for _, r in df.iterrows():

        ws.cell(row=fila+1, column=1, value=r["nombre"])
        ws.cell(row=fila+1, column=2, value=r["cantidad"])
        ws.cell(row=fila+1, column=3, value=r["fecha_cad"])

        fila += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer

# ------------------ SIDEBAR RESET ------------------

st.sidebar.header("⚙️ Configuración")

with st.sidebar.expander("🚨 Zona de Peligro"):
    confirmar_reset = st.checkbox("Confirmar que deseo borrar todo")

    if st.button("⚠️ EJECUTAR RESET TOTAL"):

        if confirmar_reset:

            c.execute("DELETE FROM captura_actual")
            c.execute("DELETE FROM base_anterior")
            c.execute("DELETE FROM historial_ventas")
            conn.commit()

            st.sidebar.success("Base limpiada")
            time.sleep(1)
            st.rerun()

        else:
            st.sidebar.error("Debes confirmar")

# ------------------ TABS ------------------

tab1, tab2, tab3 = st.tabs(["📝 Conteo", "📦 Inventario y Corte", "📊 Análisis"])

# ------------------------------------------------------------
# TAB 1
# ------------------------------------------------------------

with tab1:

    if "conteo_temp" not in st.session_state:
        st.session_state.conteo_temp = 0

    buscar = st.text_input("Buscar producto").upper()

    nombres_prev = [r[0] for r in c.execute(
        "SELECT DISTINCT nombre FROM base_anterior UNION SELECT DISTINCT nombre FROM captura_actual"
    ).fetchall()]

    sugerencias = [p for p in nombres_prev if buscar in p] if buscar else nombres_prev

    nombre_input = st.selectbox("Producto", sugerencias)

    f_cad = st.date_input("Caducidad", value=fecha_hoy_mx)

    c1, c2, c3, c4 = st.columns(4)

    with c1: st.button("+1", on_click=sumar, args=(1,))
    with c2: st.button("+5", on_click=sumar, args=(5,))
    with c3: st.button("+10", on_click=sumar, args=(10,))
    with c4: st.button("Borrar", on_click=resetear)

    st.metric("Total", st.session_state.conteo_temp)

    if st.button("➕ Registrar"):

        nombre_final = nombre_input.upper()
        cant = st.session_state.conteo_temp

        c.execute(
            "INSERT INTO captura_actual VALUES (?,?,?)",
            (nombre_final, str(f_cad), int(cant))
        )

        conn.commit()

        st.success("Registrado")
        st.session_state.conteo_temp = 0
        time.sleep(1)
        st.rerun()

# ------------------------------------------------------------
# TAB 2
# ------------------------------------------------------------

with tab2:

    st.header("Stock")

    df_stock = pd.read_sql("SELECT nombre as Producto, fecha_cad as Caducidad, cantidad as Existencia FROM base_anterior", conn)

    if df_stock.empty:

        st.info("No hay inventario")

    else:

        st.dataframe(df_stock, use_container_width=True)

        msg_stock = "🍞 INVENTARIO CHAMPLITTE\n\n"

        for _, r in df_stock.iterrows():

            msg_stock += f"{r['Producto']} | {r['Existencia']} pza | {r['Caducidad']}\n"

        link_st = f"https://wa.me/{numero_whatsapp}?text={urllib.parse.quote(msg_stock)}"

        st.link_button("Enviar a WhatsApp", link_st)

    st.divider()

    st.subheader("Exportar")

    df_export = pd.read_sql("SELECT * FROM base_anterior", conn)

    if not df_export.empty:

        csv_file = generar_csv(df_export)

        st.download_button(
            "Descargar CSV",
            csv_file,
            "inventario.csv",
            "text/csv"
        )

        excel_file = generar_excel_costa_verde(conn)

        st.download_button(
            "Descargar Excel COSTA VERDE",
            excel_file,
            "produccion.xlsx"
        )

    st.divider()

    st.header("Procesar Corte")

    if st.button("PROCESAR CORTE"):

        df_actualizado = pd.read_sql("SELECT * FROM captura_actual", conn)

        if df_actualizado.empty:

            st.warning("No hay conteo")

        else:

            c.execute("DELETE FROM base_anterior")
            c.execute("INSERT INTO base_anterior SELECT * FROM captura_actual")
            c.execute("DELETE FROM captura_actual")

            conn.commit()

            st.success("Corte procesado")
            st.balloons()

# ------------------------------------------------------------
# TAB 3
# ------------------------------------------------------------

with tab3:

    df_hist = pd.read_sql(
        "SELECT nombre as Producto, vendidos as Vendidos, fecha_corte as Fecha FROM historial_ventas",
        conn
    )

    if df_hist.empty:

        st.info("Sin historial")

    else:

        st.dataframe(df_hist)

        ventas_dia = df_hist.groupby("Fecha")["Vendidos"].sum().reset_index()

        st.line_chart(ventas_dia, x="Fecha", y="Vendidos")

        top = df_hist.groupby("Producto")["Vendidos"].sum().sort_values(ascending=False)

        if not top.empty:

            st.subheader("Producto Estrella")

            st.metric(top.index[0], f"{int(top.iloc[0])} vendidos")

            st.bar_chart(top)
